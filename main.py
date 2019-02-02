# python 3.7.*
# https://openpyxl.readthedocs.io/en/stable/tutorial.html - Open XML Writer for XLSX
# pip install openpyxl
# pip install pillow

# also note when working with huge datasets, please use numpy
# https://openpyxl.readthedocs.io/en/stable/pandas.html

from openpyxl import Workbook
from openpyxl.styles import PatternFill

# Create the workbook
wb = Workbook()
ws = wb.active

# Cell style
fill = PatternFill("solid", fgColor="00DD00")

ws1 = wb.create_sheet("Test Sheet")
ws.title = "Test XLSX Write"
c = ws['C5']
c.value = 'This is a test.'

# Apply style
c.fill = fill

# additional documentation on class implementation for reuse
# https://openpyxl.readthedocs.io/en/stable/styles.html?highlight=fill%20color

wb.save("test.xlsx")
